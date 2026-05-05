import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """
    Exports parsed bank statement data to formatted Excel files.
    Includes professional styling, headers, and auto-adjusted columns.
    """
    
    def __init__(self, output_path):
        """Initialize exporter with output file path."""
        self.output_path = output_path
        self.wb = None
        self.ws = None
    
    def export_transactions(self, transactions, summary=None):
        """
        Export transactions to Excel with professional formatting.
        
        Args:
            transactions: List of transaction dictionaries
            summary: Optional summary statistics dictionary
        """
        # Create workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Transactions"
        
        # Define headers
        headers = ['Date', 'Description', 'Amount', 'Transaction Type']
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write transaction data
        for row_num, transaction in enumerate(transactions, 2):
            date_val = transaction.get('date', '')
            desc_val = transaction.get('description', '')
            amount_val = transaction.get('amount', '')
            
            # Determine transaction type (Debit/Credit)
            trans_type = "Debit" if amount_val and amount_val.startswith('-') else "Credit"
            
            # Write cells
            self.ws.cell(row=row_num, column=1).value = date_val
            self.ws.cell(row=row_num, column=2).value = desc_val
            self.ws.cell(row=row_num, column=3).value = amount_val
            self.ws.cell(row=row_num, column=4).value = trans_type
            
            # Apply borders and alignment
            for col in range(1, 5):
                cell = self.ws.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 40
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        
        # Add summary sheet if provided
        if summary:
            self._add_summary_sheet(summary)
        
        # Save workbook
        self.wb.save(self.output_path)
        print(f"✅ Excel file created: {self.output_path}")
    
    def _add_summary_sheet(self, summary):
        """Add summary statistics sheet to workbook."""
        ws_summary = self.wb.create_sheet("Summary")
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Transactions', summary.get('total_transactions', 0)],
            ['Total Amount', f"${summary.get('total_amount', 0):.2f}"],
            ['Average Amount', f"${summary.get('average_amount', 0):.2f}"],
            ['Maximum Amount', f"${summary.get('max_amount', 0):.2f}"],
            ['Minimum Amount', f"${summary.get('min_amount', 0):.2f}"],
        ]
        
        # Write summary data
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Style header row
                if row_num == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
    
    def export_pandas_dataframe(self, df):
        """
        Export a Pandas DataFrame directly to Excel.
        
        Args:
            df: Pandas DataFrame to export
        """
        df.to_excel(self.output_path, sheet_name='Transactions', index=False)
        print(f"✅ Excel file created: {self.output_path}")
